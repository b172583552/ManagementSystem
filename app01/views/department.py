from django.shortcuts import render, redirect
from app01 import models
from django import forms
from django.core.paginator import Paginator


def department_list(request):
    queryset = models.Department.objects.all()
    no_of_items = 5
    page_object = Paginator(queryset, no_of_items)
    page_number = request.GET.get("page", "1")
    page = page_object.get_page(page_number)
    queryset = page

    return render(request, "department_list.html", {"queryset": queryset})


def department_add(request):
    if request.method == "GET":
        attributes = models.Department.objects.all()[0].__dict__
        dict = {}
        for index, attribute in enumerate(attributes):
            if index > 1:
                dict[attribute] = attributes[attribute]
        return render(request, "change.html",
                      {"name": "Department", "attributes": dict})

    title = request.POST.get("title")

    models.Department.objects.create(title=title)

    return redirect("/department/list/")


def department_delete(request):
    nid = request.GET.get("nid")
    models.Department.objects.filter(id=nid).delete()

    return redirect("/department/list/")


def department_edit(request, nid):
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        print(row_object.id, row_object.title)
        return render(request, "department_edit.html", {"row_object": row_object})

    else:
        title = request.POST.get("title")
        models.Department.objects.filter(id=nid).update(title=title)
        return redirect("/department/list")

def department_upload(request):
    from openpyxl import load_workbook
    file = request.FILES.get("department")


    wb = load_workbook(file)
    sheet = wb.worksheets[0]


    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect("/department/list/")